# -*- coding: utf-8 -*- 
import operator
import datetime
import math

from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse, resolve
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.views.decorators.cache import never_cache
from django.db.models import Q, Max, Min
from django import db
from django.conf import settings
from django.template.defaultfilters import date as tmp_date

from user_registration.func import *
from kinoinfo_folder.func import low, capit, del_separator
from api.models import FilmsName
from api.views import get_film_data, film_poster2
from base.models import *
from base.func import create_kinoafisha_button
from articles.views import pagination as pagi
from movie_online.IR import check_int_rates


def get_bg(request, site=None):
    new_view = 1
    if '/dajaxice/' in request.path or '/user/' in request.path and '/login/' not in request.path:
        new_view = 0

    btype = '2'

    price = get_adv_price(btype)

    advert_img, advert_url, advert_id = ('', '', '')

    now = datetime.datetime.now()

    if request.current_site.domain in ('kinoinfo.ru', 'kinoafisha.ru'):
        filter = {'sites': request.current_site, 'btype': btype, 'user__personinterface__money__gte': price, 'balance__gte': price, 'deleted': False, 'dtime__lt': now}
        if site:
            filter['sites'] = site
        banners = SiteBanners.objects.filter(Q(country__pk=request.current_user_country_id) | Q(country=None), **filter).values('file', 'url', 'id', 'views', 'country', 'cities__pk', 'dtime').order_by('-cities__pk')
        
        banner = None
        get_country = False
        for i in banners:
            future = i['dtime'].date() + datetime.timedelta(days=13)
            if now.date() <= future:
                # если баннер для моего города
                if i['cities__pk'] == request.current_user_city_id:
                    banner = i
                    break
                # если баннер для моей страны
                elif i['country'] == request.current_user_country_id and not i['cities__pk']:
                    banner = i
                    get_country = True
                elif not get_country and not i['country'] and not i['cities__pk']:
                    banner = i

        if banner:
            advert_img, advert_url = (banner['file'], banner['url'])
            advert_id = banner['id']
            if new_view and not request.bot:
                set_adv_view(request, banner['id'])

    else:
        bg_filter = {'country__id': request.current_user_country_id, 'city__id': request.current_user_city_id, 'site': request.current_site, 'subdomain': request.subdomain}
        bg_alt_filter = {'country__id': request.current_user_country_id, 'city': None, 'site': request.current_site, 'subdomain': request.subdomain}
        bg = Background.objects.filter(Q(**bg_filter) | Q(**bg_alt_filter)).order_by('-date_adding')
        if bg:
            advert_img, advert_url = ('%s%s' % (settings.MEDIA_URL, bg[0].image), bg[0].url)

    return {'img': advert_img, 'url': advert_url, 'id': advert_id}


@never_cache
def search(request):

    query_orig = request.GET.get('query', '')
    category = request.GET.get('filter', '')
    objs = []
    msg = ''
    element = ''
    data = {}
    count = 0
    
    if query_orig:
        
        query = low(query_orig)
        
        slug = del_separator(query)
        if len(query) > 2:
            
            # фильмы
            if category == '1':
                data = []
                tmp_query = query.decode('utf-8').encode('ascii', 'xmlcharrefreplace')
                element = 'фильмов'
                try:

                    '''
                    Получаем ID фильмов по поисковому запросу
                    upd : проверка ч/з пересечение множеств показала,
                    что такой вариант идентичен предыдущему(конкатенации двух сетов)
                    upd : убрал поиск по слагу, т.к. формирование слага происходит посредством склейки
                    всех слов из тайтла, а это приводит к нерелевантным результатам поиска
                    (см. "San tau jin zi lei saam" фильм)
                    '''

                    ids = set(list(
                        FilmsName.objects.using('afisha').filter(
                            Q(name__icontains = query),
                            status=1
                        ).values_list('film_id', flat=True)))


                    objs = FilmsName \
                        .objects \
                        .using('afisha') \
                        .only(
                            'name',
                            'film_id',
                            'film_id__description',
                            'film_id__year',
                            'film_id__genre1',
                            'film_id__genre2',
                            'film_id__imdb',
                              ) \
                        .select_related('film_id') \
                        .filter(film_id__id__in=ids, status=1, type=2) \
                        .order_by('-film_id__year','name')



                    tmp = None
                    tmp_first = []
                    relevant = []
                    nonrelevant = []

                    for i in objs:

                        '''
                        Получаем постер
                        '''

                        temp_poster = Objxres.objects.using('afisha').select_related('extresid').filter(
                            objtypeid__in=[301, 300],
                            objpkvalue=i.film_id_id)

                        posters = {}
                        for p in temp_poster:
                            if posters.get(p.objpkvalue):
                                if p.objtypeid == 301:
                                    posters[p.objpkvalue]['poster'].append(p)
                                else:
                                    posters[p.objpkvalue]['slides'].append(p)
                            else:
                                posters[p.objpkvalue] = {'poster': [], 'slides': []}
                                if p.objtypeid == 301:
                                    posters[p.objpkvalue]['poster'].append(p)
                                else:
                                    posters[p.objpkvalue]['slides'].append(p)

                        if posters and posters[i.film_id_id]['poster']:
                            i.poster = film_poster2(posters[i.film_id_id]['poster'], 'big')

                        '''
                        Получаем мету
                        '''
                        meta = Film.objects.using('afisha').get(id=i.film_id_id)

                        '''
                        Из полученных метаданных получаем наименования жанров
                        '''
                        i.genres = AfishaGenre.objects.using('afisha').filter(
                            pk__in=[meta.genre1_id, meta.genre2_id, meta.genre3_id])

                        '''
                        Из полученных метаданных получаем рейтинг по IMDB
                        грязнохак - шаблонизатор django не понимает числовых циклов с условиями
                        '''
                        if(meta.imdb):
                            i.imdb = {}
                            imdb_val = meta.imdb
                            i.imdb['range'] = range(
                                0,
                                int(
                                    math.ceil(
                                        float(
                                            imdb_val.replace(',', '.')
                                        )
                                    )
                                )
                                ,1)
                            i.imdb['value'] = meta.imdb

                        '''
                        Получаем рейтинг взвешенный по системе
                        '''
                        int_rate, show_ir, show_imdb, rotten = check_int_rates(i.film_id_id)
                        i.rating = {'rate': int_rate, 'show_ir': show_ir, 'show_imdb': show_imdb, 'rotten': rotten}

                        '''
                        upd: переработка сортировки по релевантности, признак - наличие поисковой фразы в описании
                        '''
                        tmp = i.film_id_id

                        if(tmp not in tmp_first):

                            if(
                                        (meta.description is not None)
                                    and (meta.description.find(tmp_query))
                            ):
                                relevant.append(i)
                            else:
                                nonrelevant.append(i)

                            tmp_first.append(tmp)

                    data = relevant+nonrelevant

                    count = len(tmp_first)

                    if count == 1:
                        return HttpResponseRedirect(reverse('get_film', kwargs={'id': tmp}))

                except db.backend.Database._mysql.OperationalError:
                    pass
            # персоны
            elif category == '2':
                data = {'first': [], 'middle': [], 'last': []}
                
                element = 'персон'
                
                tmp_query = query.decode('utf-8').encode('ascii', 'xmlcharrefreplace')
                
                ids1 = set(list(NamePerson.objects.exclude(person__kid=None).filter(Q(name__iexact=tmp_query) | Q(name__istartswith=slug), status__in=(1,2)).values_list('person__id', flat=True)))
                
                ids2 = set(list(NamePerson.objects.exclude(person__id__in=ids1, person__kid=None).filter(name__icontains=slug, status=2).values_list('person__id', flat=True)))
                
                ids = set(list(ids1) + list(ids2))
                
                objs1 = list(NamePerson.objects.filter(person__id__in=ids, status=1, language__id=1).order_by('name').values('person__id', 'name', 'person__kid'))
                
                result_ids = set([i['person__id'] for i in objs1])
                
                objs2 = list(NamePerson.objects.exclude(person__id__in=result_ids).filter(person__id__in=ids, status=1, language__id=2).order_by('name').values('person__id', 'name', 'person__kid'))
                
                
                tmp = None
                    
                tmp_first = []
                
                for objs in [objs1, objs2]:
                    for i in objs:
                        tmp = i['person__kid']
                        if long(i['person__id']) in ids1:
                            if low(i['name'].encode('utf-8')) == tmp_query:
                                if tmp not in tmp_first:
                                    data['first'].append((i['person__kid'], i['name']))
                                    tmp_first.append(tmp)
                            else:
                                if tmp not in tmp_first:
                                    data['middle'].append((i['person__kid'], i['name']))
                                    tmp_first.append(tmp)
                        else:
                            if tmp not in tmp_first:
                                data['last'].append((i['person__kid'], i['name']))
                                tmp_first.append(tmp)
         
                count = len(tmp_first)
                
                if count == 1:
                    return HttpResponseRedirect(reverse('get_person', kwargs={'id': tmp}))
            
            elif category == '3':
                # music
                if request.subdomain == 'music':
                    data = {'first': [], 'middle': [], 'last': [], 'artist': []}
                    
                    element = ''
                    
                    tmp_query = query.decode('utf-8').encode('ascii', 'xmlcharrefreplace')
                    
                    ids1 = set(list(Composition.objects.filter(Q(name__name__iexact=tmp_query) | Q(name__name__istartswith=slug), name__status__in=(2,5)).values_list('id', flat=True)))
                    
                    ids2 = set(list(Composition.objects.exclude(pk__in=ids1).filter(name__name__icontains=slug,  name__status__in=(2,5)).values_list('id', flat=True)))
                    
                    ids = set(list(ids1) + list(ids2))
                    
                    comp_rels = {}
                    for i in CompositionPersonRel.objects.filter(composition__pk__in=ids, type__name='исполнение').values('person', 'person__name__name', 'composition'):
                        comp_rels[i['composition']] = {'pid': i['person'], 'pname': i['person__name__name']}
                    
                    objs1 = list(CompositionName.objects.filter(composition__pk__in=ids, status=2).order_by('name').values('composition__id', 'name'))
                    
                    tmp = None
                        
                    tmp_first = []
                    
                    for i in objs1:
                        tmp = i['composition__id']
                        if long(i['composition__id']) in ids1:
                            if low(i['name'].encode('utf-8')) == query:
                                if tmp not in tmp_first:
                                    artsit = comp_rels.get(i['composition__id'])
                                    data['first'].append((i['composition__id'], i['name'], artsit))
                                    tmp_first.append(tmp)
                            else:
                                if tmp not in tmp_first:
                                    artsit = comp_rels.get(i['composition__id'])
                                    data['middle'].append((i['composition__id'], i['name'], artsit))
                                    tmp_first.append(tmp)
                        else:
                            if tmp not in tmp_first:
                                artsit = comp_rels.get(i['composition__id'])
                                data['last'].append((i['composition__id'], i['name'], artsit))
                                tmp_first.append(tmp)
             
                    count = len(tmp_first)

                    artists = Person.objects.filter(Q(name__name__iexact=tmp_query) | Q(name__name__istartswith=slug) | Q(name__name__icontains=slug), artist=True, name__status=4).values('id', 'name__name')

                    data['artist'] = artists
                    
                    count += artists.count()
                # Кинотеатры
                else:

                    element = 'кинотеатров'
                    
                    tmp_query = query.decode('utf-8').encode('ascii', 'xmlcharrefreplace')
                    
                    ids1 = set(list(Organization.objects.exclude(kid=None).filter(Q(name__iexact=tmp_query) | Q(name__istartswith=slug)).values_list('id', flat=True)))
                    ids2 = set(list(Organization.objects.exclude(Q(id__in=ids1) | Q(kid=None)).filter(name__icontains=slug).values_list('id', flat=True)))
                    
                    ids = set(list(ids1) + list(ids2))

                    buildings = list(Building.objects.filter(organization__id__in=ids, city__name__status=1).values('city', 'city__name__name', 'city__country', 'city__country__name', 'city__kid', 'organization', 'organization__name', 'organization__uni_slug'))

                    
                    cities = {}
                    count = 0
                    tmp = None
                    for i in buildings:
                        count += 1
                        tmp = i['organization__uni_slug']
                        if not cities.get(i['city']):
                            cities[i['city']] = {
                                'id': i['city'], 
                                'name': i['city__name__name'], 
                                'country': i['city__country__name'], 
                                'cinemas': []
                            }

                        cities[i['city']]['cinemas'].append({'name': i['organization__name'], 'slug': i['organization__uni_slug']})

                    cities = sorted(cities.values(), key=operator.itemgetter('name'))


                    data = {}
                    for i in cities:
                        if not data.get(i['country']):
                            data[i['country']] = []
                        data[i['country']].append(i)
                    
                    if count == 1:
                        return HttpResponseRedirect(reverse('organization_cinema', kwargs={'id': tmp}))

        else:
            msg = 'Слишком короткий запрос'
    
    return render_to_response('search_result.html', {'objs': data, 'element': element, 'msg': msg, 'srch_category': category, 'query': query_orig, 'count': count},  context_instance=RequestContext(request))
    
    
@never_cache
def kinoafisha_old_button(request, id):
    if request.current_site.domain == 'kinoafisha.ru':
        try:
            film = Film.objects.using('afisha').get(idalldvd=int(id))
        except Film.DoesNotExist:
            raise Http404
            
        img = create_kinoafisha_button(film, id)
        
        with open(img['output_path'], 'rb') as f:
            return HttpResponse(f.read(), mimetype="image/png")
            
    raise Http404


@never_cache
def kinoafisha_button(request):
    from urlparse import urlparse, parse_qs
    from release_parser.imdb import create_film_by_imdb_id

    warning = ''
    img = ()
    
    if request.POST:
        url = request.POST.get('url','').strip()
        if url:
            imdb_id = None
            film_id = None
            if 'imdb.com/title/tt' in url:
                xurl = url.split('imdb.com/title/tt')[1].split('/')[0].replace('/','')
                xurl = xurl.split('?')[0]
                imdb_id = int(xurl)
            elif 'kinoafisha.ru/' in url:
                xurl = urlparse(url)
                if xurl.query:
                    params = parse_qs(xurl.query)
                    film_id = params.get('id1')
                    film_id = film_id[0] if film_id else None
                else:
                    film_id = xurl.path.replace('/film/','').replace('/','')
                    
                try:
                    film_id = int(film_id)
                except ValueError:
                    film_id = None

            if imdb_id or film_id:
                if imdb_id:
                    try:
                        film = Film.objects.using('afisha').get(idalldvd=imdb_id)
                    except Film.DoesNotExist:
                        film = create_film_by_imdb_id(imdb_id)
                else:
                    try:
                        film = Film.objects.using('afisha').get(pk=film_id)
                    except Film.DoesNotExist:
                        film = None
                    
                if film:
                    img = create_kinoafisha_button(film, imdb_id)
                else:
                    warning = 'Фильм не найден в базе'
            else:
                warning = 'Неверный URL'
        else:
            warning = 'Вы не ввели URL'
    return render_to_response('kinoafisha/button.html', {'warning': warning, 'img': img},  context_instance=RequestContext(request))
    


def disable_adv_bg_func():
    price = 1
    now = datetime.datetime.now()

    objs = SiteBanners.objects.select_related('user', 'user__personinterface').filter(btype='0', user__personinterface__money__gte=price, balance__gte=price)

    for i in objs:
        interface = i.user.personinterface
        if interface.money >= price and i.balance >= price:
            to = now + datetime.timedelta(days=1)
            if i.bg_disable_dtime_to.date() != to.date():
                interface.money -= price
                interface.save()
                i.bg_disable_dtime_to = to
                i.balance -= price
                i.spent += price
                i.save()

@never_cache
def get_torrent(request, id):

    price = 1

    try:
        torrent = Torrents.objects.exclude(path=None).get(pk=id)
    except Torrents.DoesNotExist:
        raise Http404

    try:
        access = TorrentsUsers.objects.get(torrent=torrent, profile=request.profile)
    except TorrentsUsers.DoesNotExist:
        access = False

    interface = request.profile.personinterface

    if not access:
        if interface.money >= price:

            interface.money -= price
            interface.save()

            access, created = TorrentsUsers.objects.get_or_create(
                torrent = torrent, 
                profile = request.profile,
                defaults = {
                    'torrent': torrent, 
                    'profile': request.profile,
                })

    if access:
        full_path = '%s%s' % (settings.MEDIA_ROOT, torrent.path)
        filename = full_path.split('/')[-1]
        with open(full_path, 'r') as f:
            response = HttpResponse(f.read(), mimetype='application/x-bittorrent')
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            access.got = True
            access.save()
            return response

    raise Http404


@never_cache
def new_torrents(request):
    from slideblok.views import releasedata
    from movie_online.IR import check_int_rates_inlist

    price = 100

    interface = request.profile.personinterface

    year, genre, country, rate = (None, None, None, None)
    rate_filter = [u'ВСЕ',]
    year_filter = [u'ВСЕ',]
    genre_filter = {0: {'id': 0, 'name': u'ВСЕ'}}
    country_filter = {0: {'id': 0, 'name': u'ВСЕ'}}
    
    set_filter = False
    access = False
    if interface.money >= price or request.user.is_superuser:
        access = True

        torrents_kids = list(Torrents.objects.exclude(path=None).distinct('film').values('id', 'film'))

        kids = [i['film'] for i in torrents_kids]

        opinions = {'good': [], 'bad': []}
        for i in NewsFilms.objects.filter(kid__in=kids, message__visible=True, message__autor=request.profile, rate_1__gt=0).values('kid', 'rate'):
            if i['rate'] >= 4:
                opinions['good'].append(i['kid'])
            else:
                opinions['bad'].append(i['kid'])

        kids = set(kids) - set(opinions['bad'])

        exist_kids = []
        for i in list(Film.objects.using('afisha').filter(pk__in=kids, date__gte=datetime.date(1900,1,1)).values('year', 'genre1', 'genre2', 'genre3', 'genre1__name', 'genre2__name', 'genre3__name', 'country', 'country2', 'country__name', 'country2__name', 'id')):
            if i['year']:
                year_filter.append(int(i['year']))

            for j in ((i['genre1'], i['genre1__name']), (i['genre2'], i['genre2__name']), (i['genre3'], i['genre3__name'])):
                if j[0] and not genre_filter.get(j[0]):
                    genre_filter[int(j[0])] = {'id': int(j[0]), 'name': j[1]}

            for j in ((i['country'], i['country__name']), (i['country2'], i['country2__name'])):
                if j[0] and not country_filter.get(j[0]):
                    country_filter[int(j[0])] = {'id': int(j[0]), 'name': j[1]}

            exist_kids.append(i['id'])

        year_filter = sorted(set(year_filter), reverse=True)

        rates = check_int_rates_inlist(exist_kids)
        
        rates_tmp = set([i['int_rate'] for i in rates.values() if i['int_rate']])

        for i in rates_tmp:
            rate_filter.append(i)

        
        if request.POST:
            if 'filter' in request.POST:
                try: year = int(request.POST.get('year'))
                except ValueError: pass
                try: genre = int(request.POST.get('genre'))
                except ValueError: pass
                try: country = int(request.POST.get('country'))
                except ValueError: pass
                try: rate = int(request.POST.get('rate'))
                except ValueError: pass


        if year not in year_filter:
            year = year_filter[0]
        if genre not in genre_filter:
            genre = genre_filter[0]['id']
        if country not in country_filter:
            country = country_filter[0]['id']
        if rate not in rate_filter:
            rate = rate_filter[0]

        if year != u'ВСЕ' or rate != u'ВСЕ' or genre != 0 or country != 0:
            set_filter = True

        genre_filter = sorted(genre_filter.values(), key=operator.itemgetter('name'))
        country_filter = sorted(country_filter.values(), key=operator.itemgetter('name'))
        
        filter = {'pk__in': kids, 'date__gte': datetime.date(1900,1,1)}
        
        if year != u'ВСЕ':
            filter['year__exact'] = year

        queries = []
        if genre:
            for q in [Q(genre1__pk=genre), Q(genre2__pk=genre), Q(genre3__pk=genre)]:
                queries.append(q)
        if country:
            for q in [Q(country__pk=country), Q(country2__pk=country)]:
                queries.append(q)

        if queries:
            query = queries.pop()
            for item in queries:
                query |= item

            torrents_kids = list(Film.objects.using('afisha').filter(query, **filter).values('pk', 'date'))
        else:
            torrents_kids = list(Film.objects.using('afisha').filter(**filter).values('pk', 'date'))

        kids = [i['pk'] for i in torrents_kids]

        tusers = list(TorrentsUsers.objects.filter(profile=request.profile, torrent__film__in=kids, got=True).values_list('torrent__film', flat=True))


        for i in torrents_kids:
            if not i['date']:
                i['date'] = datetime.datetime(3000,1,1)

        torrents_kids = sorted(torrents_kids, key=operator.itemgetter('date'), reverse=True)

        page = request.GET.get('page')
        try:
            page = int(page)
        except (ValueError, TypeError):
            page = 1
        
        p, page = pagi(page, torrents_kids, 20)

        torrents_films_dict = {}
        for i in p.object_list:
            torrents_films_dict[i['pk']] = {}

        data = []
        
        for i in releasedata(torrents_films_dict, {}, persons=True, likes=True, trailers=True, reviews=True, poster_size='big'):
            i['got'] = True if i['id'] in tusers else False
            if not i['release_date']:
                i['release_date'] = datetime.datetime(3000,1,1)

            if i['id'] in opinions['good']:
                i['opinion'] = True

            if rate == u'ВСЕ' or int(rate) == int(i['rate']):
                data.append(i)

        data = sorted(data, key=operator.itemgetter('release_date'), reverse=True)
    else:
        p = None
        page = 1
        data = []

    tmplt = 'kinoafisha/new_torrents.html'

    url_name = resolve(request.path_info).url_name

    return render_to_response(tmplt, {'data': data, 'page': page, 'p': p, 'url_name': url_name, 'year_filter': year_filter, 'genre_filter': genre_filter, 'country_filter': country_filter, 'year': year, 'genre': genre, 'country': country, 'access': access, 'rate_filter': rate_filter, 'rate': rate, 'set_filter': set_filter}, context_instance=RequestContext(request))


@never_cache
def best_torrents(request):
    from slideblok.views import releasedata
    
    price = 100

    interface = request.profile.personinterface

    access = False
    if interface.money >= price:
        access = True

    else:
        pass

    tmplt = 'kinoafisha/best_torrents.html'

    url_name = resolve(request.path_info).url_name
    
    return render_to_response(tmplt, {'data': data, 'page': page, 'p': p, 'url_name': url_name, 'year_filter': year_filter, 'genre_filter': genre_filter, 'country_filter': country_filter, 'year': year, 'genre': genre, 'country': country, 'access': access}, context_instance=RequestContext(request))


@only_superuser
@never_cache
def torrents_listing(request, source):
    from slideblok.views import releasedata
    from news.views import cut_description
    from movie_online.IR import check_int_rates_inlist
    from release_parser.func import get_file_modify_datetime

    source = ImportSources.objects.get(pk=source)

    today = datetime.datetime.now().date()
    yesterday = today - datetime.timedelta(days=1)

    file_add = 'xml' if source.url == 'http://rutracker.org/' else 'html'
    file_path = '%s__%s__films.xml' % (file_add, source.dump)
    parser_time = get_file_modify_datetime(settings.SUCCESS_LOG_PATH, file_path)
    parser_time = parser_time + datetime.timedelta(hours=3)

    release, new, rate = (None, None, None)

    release_filter = {
        0: {'id': 0, 'name': u'ВСЕ'},
        1: {'id': 1, 'name': u'Кинопрокат'},
        2: {'id': 2, 'name': u'Прочие'},
    }

    new_filter = {
        0: {'id': 0, 'name': u'ВСЕ'},
        1: {'id': 1, 'name': u'Новые'},
    }

    rate_filter = [u'ВСЕ',]


    if request.POST:
        if 'checker' in request.POST:
            checker = request.POST.getlist('checker')
            if checker:
                SourceFilms.objects.filter(source_obj=source, pk__in=checker).update(rel_ignore=True)
            return HttpResponseRedirect(reverse('torrents_listing', kwargs={'source': source.pk}))
        else:
            release = request.POST.get('release')
            new = request.POST.get('new')
            rate = request.POST.get('rate')
    

    kids = list(SourceFilms.objects.filter(source_obj=source, rel_ignore=False).values_list('kid', flat=True))

    rates = check_int_rates_inlist(kids)
    rates_tmp = set([i['int_rate'] for i in rates.values() if i['int_rate']])
    for i in rates_tmp:
        rate_filter.append(i)

    if rate:
        try:
            rate = int(rate)
        except ValueError:
            rate = u'ВСЕ'
    if release:
        release = int(release)
    if new:
        new = int(new)
    sess_filter = request.session.get('torrents_listing_filter',{})

    if rate == None and release == None and new == None:
        sess_filter = request.session.get('torrents_listing_filter',{})
        if sess_filter:
            rate = sess_filter['rate']
            release = sess_filter['release']
            new = sess_filter['new']
        else:
            new = 1


    if not new:
        source_films = SourceFilms.objects.filter(source_obj=source, rel_ignore=False)
    elif new == 1:
        source_films = SourceFilms.objects.filter(source_obj=source, extra='new', rel_ignore=False)


    tmp_date = datetime.datetime(3000,1,1)

    films = {}
    for i in source_films:
        new = True if i.extra == 'new' else False
        films[i.kid] = {'source_id': i.source_id, 'new': new, 'pk': i.id, 'release': tmp_date, 'kid': i.kid}

    for i in list(Film.objects.using('afisha').filter(pk__in=films.keys(), date__gte=datetime.datetime(1900,1,1)).values('id', 'date')):
        films[i['id']]['release'] = i['date']

    films_sorted = sorted(films.values(), key=operator.itemgetter('release'), reverse=True)

    page = request.GET.get('page')
    try:
        page = int(page)
    except (ValueError, TypeError):
        page = 1
    
    p, page = pagi(page, films_sorted, 100)

    films_ids = {}
    for i in p.object_list:
        films_ids[i['kid']] = {}

    torrents = {}
    for i in Torrents.objects.filter(film__in=films_ids.keys()).exclude(path=None):
        if not torrents.get(i.film):
            torrents[i.film] = {'0': None, '1': None, '2': None}
        q = i.quality_avg if i.quality_avg else '1'
        torrents[i.film][q] = i

    data_tmp = releasedata(films_ids, {}, persons=False, likes=False, trailers=False, reviews=False, poster_size='small')

    upd = []

    data = []
    for i in data_tmp:      
        torrent = torrents.get(i['id'])
        f = films.get(i['id'])
        txt_cut = cut_description(i['descript'], True, 150)
        i['descript_cut'] = txt_cut
        i['descript'] = ''
        i['new'] = f['new']
        i['source_id'] = f['pk']
        i['torrent'] = torrent

        if not i['release_date']:
            i['release_date'] = datetime.datetime(3000,1,1)

        if source.url == 'http://cinemate.cc/':
            i['source_url'] = u'http://cinemate.cc/movie/%s/' % f['source_id']
        elif source.url == 'http://rutracker.org/':
            i['source_url'] = u'http://rutracker.org/forum/viewtopic.php?t=%s' % f['source_id']

        if f['new']:
            upd.append(f['pk'])
        
        next = True

        if release == 1 and i['release_date'].year == 3000:
            next = False
        elif release == 2 and i['release_date'].year < 3000:
            next = False
        
        if next:
            if rate:
                if rate == u'ВСЕ' or rate == int(i['rate']):
                    data.append(i)
            else:
                data.append(i)

    films = sorted(data, key=operator.itemgetter('release_date'), reverse=True)

    SourceFilms.objects.filter(pk__in=upd).update(extra=None)

    request.session['torrents_listing_filter'] = {'rate': rate, 'new': new, 'release': release}

    return render_to_response('kinoafisha/torrents_listing.html', {'data': films, 'source': source, 'p': p, 'page': page, 'release_filter': release_filter.values(), 'release': release, 'new_filter': new_filter.values(), 'new': new, 'rate_filter': rate_filter, 'rate': rate, 'parser_time': parser_time, 'today': today, 'yesterday': yesterday}, context_instance=RequestContext(request))


@never_cache
def booking(request):

    timer = time.time()

    today = datetime.datetime.today().date()

    def get_timeline():
        next_day = today + datetime.timedelta(days=1)
        timeline = [{'time': datetime.datetime(today.year, today.month, today.day, 9, 0), 'minutes': False}]
        while True:
            time_to = timeline[-1]['time'] + datetime.timedelta(minutes=15)
            if time_to.date() == next_day:
                break
            minutes = False if time_to.minute == 0 else True
            timeline.append({'time': time_to, 'minutes': minutes})
        return timeline


    today = datetime.date.today()
    date_from = today

    # Получаем релизы по стране
    country = 2

    if request.POST:
        country = int(request.POST.get('country', 2))
        date_from = request.POST.get('date_from', str(date_from))
        date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d')

    releases = []

    if country == 2:
        releases = list(Film.objects.using('afisha').filter(date__gte=today).order_by('date').values('id', 'date'))
    elif country == 43:
        for i in UkrainianReleases.objects.filter(release__gte=today).order_by('release'):
            releases.append({'id': i.kid, 'date': i.release})

    kids = set([i['id'] for i in releases])

    names = {}
    for i in FilmsName.objects.using('afisha').filter(film_id__pk__in=kids, status=1, type__in=(1, 2)):
        if not names.get(i.film_id_id):
            names[i.film_id_id] = {}
        names[i.film_id_id][i.type] = i.name

    for i in releases:
        name = names.get(i['id'])
        i['name'] = name[2] if name.get(2) else name[1]


    # Проверяем права
    if request.user.is_superuser:
        access = True
    else:
        access = request.user.groups.filter(name='Букер').exists()


    
    articles = []
    days_range = []
    halls = {}
    data = {}

    if access:

        # Получаем заметки букера
        try:
            osm = OrgSubMenu.objects.get(page_type='1', booker_profile=request.profile)
        except OrgSubMenu.DoesNotExist:
            pass
        else:
            articles = osm.news.all().order_by('-dtime')


        # все кинотеатры из настроек букера
        cinemas = {}
        cinemas_kids = []
        for i in Cinema.objects.filter(bookingsettings__profile=request.profile).distinct('pk').values('id', 'code', 'bookercinemas__permission'):
            cinemas[i['id']] = {'id': i['id'], 'kid': i['code'], 'name': '', 'access': i['bookercinemas__permission']}
            cinemas_kids.append(i['code'])

        for i in list(NameCinema.objects.filter(cinema__code__in=cinemas_kids, status=1).values('name', 'cinema')):
            cinemas[i['cinema']]['name'] = i['name']

        # залы
        for i in list(NameHall.objects.filter(status=1, hall__cinema__code__in=cinemas_kids).distinct('hall').values('name', 'hall', 'hall__cinema')):
            cinema = cinemas.get(i['hall__cinema'])
            halls[i['hall']] = {'hall_id': i['hall'], 'cinema': cinema, 'cinema_name': cinema['name'], 'name': i['name'], 'dates': {}, 'date_range': {}}
            

        # диапозон дат сеансов
        days = BookingSchedules.objects.filter(hall__cinema__code__in=cinemas_kids, dtime__gte=today).values_list('dtime', flat=True)
        days_range = list(set([i.date() for i in days]))
        days_range.sort()

        # сеансы от ... и далее
        date_to = date_from + datetime.timedelta(days=1)
        schedules = BookingSchedules.objects.select_related('hall').filter(hall__cinema__code__in=cinemas_kids, dtime__gte=date_from)


        # фильмы
        schedules_ids = [i.id for i in schedules]
        source_films_data = list(SourceFilms.objects.filter(bookingschedules__pk__in=schedules_ids).values('kid', 'bookingschedules'))
        source_kids = set([i['kid'] for i in source_films_data])

        for i in FilmsName.objects.using('afisha').filter(film_id__pk__in=source_kids, status=1, type__in=(1, 2)).exclude(film_id__id__in=names.keys()):
            if not names.get(i.film_id_id):
                names[i.film_id_id] = {}
            names[i.film_id_id][i.type] = i.name

        
        source_films = {}
        for i in source_films_data:
            if not source_films.get(i['bookingschedules']):
                source_films[i['bookingschedules']] = []
            name = names.get(i['kid'])
            name = name[2] if name.get(2) else name[1]
            source_films[i['bookingschedules']].append({'kid': i['kid'], 'name': name})


        # удобно складываем все данные
        for i in schedules:
            if not data.get(i.hall_id):
                #cinema = cinemas.get(i.hall.cinema_id)
                hall = halls.get(i.hall_id)
                data[i.hall_id] = {'hall_id': i.hall_id, 'cinema': hall['cinema'], 'name': hall['name'], 'dates': {}, 'date_range': {}}
            films = source_films.get(i.id)
            
            if not data[i.hall_id]['dates'].get(i.dtime.date()):
                data[i.hall_id]['dates'][i.dtime.date()] = {'date': i.dtime.date(), 'ids': [], 'times': []}

            data[i.hall_id]['dates'][i.dtime.date()]['times'].append({'time': i.dtime, 'films': films, 'tmp': i.temp, 'id': i.id})
            data[i.hall_id]['dates'][i.dtime.date()]['ids'].append('%s%s' % (i.dtime.time(), films))
            data[i.hall_id]['dates'][i.dtime.date()]['ids'].sort()

        
        # формируем диапозон дат с одинкаовыми сеансами
        for i in data.values():
            i['dates'] = sorted(i['dates'].values(), key=operator.itemgetter('date'))

            old_date_times = []
            sch_date_from = None
            sch_date_to = None
            for j in i['dates']:
                set_dates = True
                if old_date_times:
                    if old_date_times == j['ids']:
                        sch_date_to = j['date']
                        set_dates = False
                
                if set_dates:
                    sch_date_from = j['date']
                    sch_date_to = j['date']

                old_date_times = j['ids']

                if not i['date_range'].get(sch_date_from):
                    i['date_range'][sch_date_from] = {'from': sch_date_from, 'to': None, 'times': j['times']}
                i['date_range'][sch_date_from]['to'] = sch_date_to

        for i in data.values():
            i['date_range'] = sorted(i['date_range'].values(), key=operator.itemgetter('from'))
            for j in i['date_range']:
                j['times'] = sorted(j['times'], key=operator.itemgetter('time'))


    halls = sorted(halls.values(), key=operator.itemgetter('cinema_name'))
    
    timeline = get_timeline()
    url_name = resolve(request.path_info).url_name
    try:
        date_from = date_from.date()
    except AttributeError:
        pass

    timer = '%5.2f' % (time.time()-timer)
    return render_to_response('kinoafisha/booking.html', {'url_name': url_name, 'timeline': timeline, 'country': country, 'releases': releases, 'articles': articles, 'access': access, 'days_range': days_range, 'data': data.values(), 'date_from': date_from, 'timer': timer, 'cached_page': '', 'halls': halls}, context_instance=RequestContext(request))



@never_cache
def booking_article_add(request):
    from letsgetrhythm.views import view_func
    if request.POST:

        if request.user.is_superuser:
            access = True
        else:
            access = request.user.groups.filter(name='Букер').exists()

        articles = []
        if access:
            osm, created = OrgSubMenu.objects.get_or_create(
                page_type = '1',
                booker_profile = request.profile,
                defaults = {
                    'name': 'Booking заметки',
                    'page_type': '1',
                    'booker_profile': request.profile,
                })
            
            data = view_func(request, osm.id, None, 'booking', access)

            return HttpResponseRedirect(reverse('booking'))

    raise Http404




@never_cache
def booking_get_excel_doc(request):
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import Font

    if request.user.is_superuser:
        access = True
    else:
        access = request.user.groups.filter(name='Букер').exists()

    data = {}
    halls = {}
    if access and request.POST:
        '''
        halls_ids = [3552, 1970, 2539]
        date_from = datetime.date(2016, 4, 25)
        date_to = datetime.date(2016, 5, 5)
        date_to = date_to + datetime.timedelta(days=1)
        print_temp = False
        all_dates = False
        '''

        halls_ids = request.POST.getlist('ex_halls')
        date_from = request.POST.get('ex_date_from')
        date_to = request.POST.get('ex_date_to')
        all_dates = request.POST.get('ex_date_all')
        print_temp = request.POST.get('ex_temp')

        date_from = datetime.datetime.strptime(str(date_from), '%Y-%m-%d')
        date_to = datetime.datetime.strptime(str(date_to), '%Y-%m-%d')
        all_dates = True if all_dates else False
        print_temp = True if print_temp else False

        # все кинотеатры из настроек букера
        cinemas = {}
        cinemas_kids = []
        for i in Cinema.objects.filter(bookingsettings__profile=request.profile, bookercinemas__permission='1').distinct('pk').values('id', 'code', 'bookercinemas__permission'):
            cinemas[i['id']] = {'id': i['id'], 'kid': i['code'], 'name': '', 'access': i['bookercinemas__permission']}
            cinemas_kids.append(i['code'])

        for i in list(NameCinema.objects.filter(cinema__code__in=cinemas_kids, status=1).values('name', 'cinema')):
            cinemas[i['cinema']]['name'] = i['name']

        # залы
        for i in list(NameHall.objects.filter(status=1, hall__cinema__code__in=cinemas_kids).distinct('hall').values('name', 'hall', 'hall__cinema')):
            cinema = cinemas.get(i['hall__cinema'])
            halls[i['hall']] = {'hall_id': i['hall'], 'cinema': cinema, 'cinema_name': cinema['name'], 'name': i['name'], 'dates': {}, 'date_range': {}}

        filter = {'hall__id__in': halls_ids, 'temp': print_temp}
        if all_dates:
            filter['dtime__gte'] = datetime.date.today()
        else:
            filter['dtime__gte'] = date_from
            filter['dtime__lt'] = date_to

        schedules = BookingSchedules.objects.select_related('hall').filter(**filter)

        schedules_ids = [i.id for i in schedules]
        source_films_data = list(SourceFilms.objects.filter(bookingschedules__pk__in=schedules_ids).values('kid', 'bookingschedules'))
        source_kids = set([i['kid'] for i in source_films_data])

        names = {}
        for i in FilmsName.objects.using('afisha').filter(film_id__pk__in=source_kids, status=1, type__in=(1, 2)).exclude(film_id__id__in=names.keys()):
            if not names.get(i.film_id_id):
                names[i.film_id_id] = {}
            names[i.film_id_id][i.type] = i.name

        # фильмы
        source_films = {}
        for i in source_films_data:
            if not source_films.get(i['bookingschedules']):
                source_films[i['bookingschedules']] = []
            name = names.get(i['kid'])
            name = name[2] if name.get(2) else name[1]
            source_films[i['bookingschedules']].append({'kid': i['kid'], 'name': name})

        # удобно складываем все данные
        for i in schedules:
            if not data.get(i.hall_id):
                #cinema = cinemas.get(i.hall.cinema_id)
                hall = halls.get(i.hall_id)
                data[i.hall_id] = {'hall_id': i.hall_id, 'cinema': hall['cinema'], 'name': hall['name'], 'dates': {}, 'date_range': {}}
            films = source_films.get(i.id)
            
            if not data[i.hall_id]['dates'].get(i.dtime.date()):
                data[i.hall_id]['dates'][i.dtime.date()] = {'date': i.dtime.date(), 'ids': [], 'times': []}

            data[i.hall_id]['dates'][i.dtime.date()]['times'].append({'time': i.dtime, 'films': films, 'tmp': i.temp, 'id': i.id})
            data[i.hall_id]['dates'][i.dtime.date()]['ids'].append('%s%s' % (i.dtime.time(), films))
            data[i.hall_id]['dates'][i.dtime.date()]['ids'].sort()

        # формируем диапозон дат с одинкаовыми сеансами
        for i in data.values():
            i['dates'] = sorted(i['dates'].values(), key=operator.itemgetter('date'))

            old_date_times = []
            sch_date_from = None
            sch_date_to = None
            for j in i['dates']:
                set_dates = True
                if old_date_times:
                    if old_date_times == j['ids']:
                        sch_date_to = j['date']
                        set_dates = False
                
                if set_dates:
                    sch_date_from = j['date']
                    sch_date_to = j['date']

                old_date_times = j['ids']

                if not i['date_range'].get(sch_date_from):
                    i['date_range'][sch_date_from] = {'from': sch_date_from, 'to': None, 'times': j['times']}
                i['date_range'][sch_date_from]['to'] = sch_date_to


        wb = Workbook()
        
        for index, i in enumerate(data.values()):
            if index == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet()

            ws.title = u'%s, %s' % (i['cinema']['name'][:13], i['name'][:13])

            count = 0
            for j in i['date_range'].values():
                count += len(j['times']) + 1

            last_row = 0
            for j in sorted(i['date_range'].values(), key=operator.itemgetter('from')):
                last_row += 1

                cells = 'A%s:B%s' % (last_row, last_row)

                ws.merge_cells(cells)
                ws["A%s" % last_row].font = Font(name='Calibri', size=12, bold=True)
                ws["A%s" % last_row] = '%s - %s' % (j['from'], j['to'])

                j['times'] = sorted(j['times'], key=operator.itemgetter('time'))

                for row in range(len(j['times'])):
                    last_row += 1

                    t = tmp_date(j['times'][row]['time'], 'H:i')
                    t = t.encode('utf-8')
                    tfilms = ''
                    for f in j['times'][row]['films']:
                        if tfilms:
                            tfilms += u', '
                        tfilms += f['name']
                    
                    for ind, col in enumerate((t, tfilms)):
                        ws.cell(column=ind+1, row=last_row, value=col)

        file_name = 'booking_schedules_%s.xlsx' % request.user.id
        file_path = '%s/%s' % (settings.BOOKING_EXCEL_PATH, file_name)

        wb.save(filename=file_path)

        with open(file_path, 'r') as f:
            file = f.read()

        response = HttpResponse(file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        return response

        '''
        doc = ''

        for i in data.values():
            doc += u'<b>%s / %s</b><br />' % (i['cinema']['name'][:13], i['name'][:13])
            
            for j in sorted(i['date_range'].values(), key=operator.itemgetter('from')):

                doc += u'- - - - - %s - %s<br />' % (j['from'], j['to'])

                for t in sorted(j['times'], key=operator.itemgetter('time')):

                    doc += u'%s, ' % t['time'].time()
                    for f in t['films']:
                        doc += u'%s, ' % f['name']
                    doc += u'<br />'

        return HttpResponse(str(doc.encode('utf-8')))
        '''

    raise Http404